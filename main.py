"""Backend de licencias/pagos/correo de PGIRASA Tools.

Custodia las llaves sensibles (service_role de Supabase, llaves Wompi y SMTP)
para que NUNCA viajen en la app de escritorio. La app llama a estos endpoints
con una clave compartida (X-App-Key).

Desplegar en Coolify (ver README.md).
"""
import base64
import hashlib
import hmac
import os
import smtplib
import threading
import time
import urllib.parse
from datetime import datetime, timedelta, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import logging

import bcrypt
import requests
from fastapi import FastAPI, Header, HTTPException, Request
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel

# ── Configuración (variables de entorno en Coolify) ──────────────────────────
APP_API_KEY        = os.getenv("APP_API_KEY", "")
# Clave EXCLUSIVA de la app móvil: solo habilita las rutas /movil/*.
# Si se filtra (al descompilar el APK), NO da acceso a licencias/pagos/sync.
MOVIL_API_KEY      = os.getenv("MOVIL_API_KEY", "")
SUPABASE_URL       = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
WOMPI_PUBLIC_KEY   = os.getenv("WOMPI_PUBLIC_KEY", "")
WOMPI_INTEGRITY_KEY = os.getenv("WOMPI_INTEGRITY_KEY", "")
WOMPI_PRIVATE_KEY  = os.getenv("WOMPI_PRIVATE_KEY", "")
SMTP_REMITENTE     = os.getenv("SMTP_REMITENTE", "")
SMTP_APP_PWD       = os.getenv("SMTP_APP_PWD", "")
SMTP_HOST          = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT          = int(os.getenv("SMTP_PORT", "587"))
PAGO_MONTO         = int(os.getenv("PAGO_MONTO", "15000"))
CODIGO_TTL_MIN     = 15

app = FastAPI(title="PGIRASA Backend", version="1.0.0")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("pgirasa")


@app.exception_handler(Exception)
async def _errores_no_controlados(request: Request, exc: Exception):
    """En vez de un 500 mudo, registra el traceback y devuelve el detalle real.

    Así, ante un fallo (p. ej. una escritura a Supabase), la app muestra la causa
    y queda en los logs de Coolify para diagnosticar.
    """
    logger.exception("Error no controlado en %s %s", request.method, request.url.path)
    detalle = f"{type(exc).__name__}: {exc}"
    resp = getattr(exc, "response", None)
    if resp is not None:
        try:
            detalle += f" | Supabase: {resp.text[:300]}"
        except Exception:
            pass
    return JSONResponse(status_code=500, content={"detail": detalle})


# ── Seguridad: clave compartida ──────────────────────────────────────────────
def _auth(x_app_key: str | None):
    """Clave principal: rutas sensibles (licencias, pagos, correo, sync)."""
    if not APP_API_KEY or x_app_key != APP_API_KEY:
        raise HTTPException(status_code=401, detail="No autorizado.")


def _auth_movil(x_app_key: str | None):
    """Rutas /movil/*: acepta la clave de móvil (limitada) o la principal.

    Así la app móvil puede usar una clave que NO abre los endpoints sensibles.
    Es retrocompatible: mientras no se configure MOVIL_API_KEY, sigue valiendo
    la clave principal que la app ya enviaba.
    """
    validas = {k for k in (APP_API_KEY, MOVIL_API_KEY) if k}
    if not validas or x_app_key not in validas:
        raise HTTPException(status_code=401, detail="No autorizado.")


# ── Freno a fuerza bruta en el login (en memoria) ────────────────────────────
_LOGIN_LOCK = threading.Lock()
_LOGIN_INTENTOS = {}      # usuario -> [timestamps de fallos recientes]
_LOGIN_MAX = 5            # fallos permitidos
_LOGIN_VENTANA = 300      # segundos (5 minutos)


def _login_bloqueado(usuario):
    ahora = time.time()
    with _LOGIN_LOCK:
        intentos = [x for x in _LOGIN_INTENTOS.get(usuario, []) if ahora - x < _LOGIN_VENTANA]
        _LOGIN_INTENTOS[usuario] = intentos
        return len(intentos) >= _LOGIN_MAX


def _login_fallo(usuario):
    with _LOGIN_LOCK:
        _LOGIN_INTENTOS.setdefault(usuario, []).append(time.time())


def _login_ok(usuario):
    with _LOGIN_LOCK:
        _LOGIN_INTENTOS.pop(usuario, None)


# ── Cliente Supabase (service_role) ──────────────────────────────────────────
def _sb_headers(extra=None):
    h = {"apikey": SUPABASE_SERVICE_KEY,
         "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
         "Content-Type": "application/json"}
    if extra:
        h.update(extra)
    return h


# Sesión HTTP persistente: reutiliza conexiones (keep-alive) hacia Supabase,
# evitando el handshake TCP en cada llamada. Acelera login y dashboard.
_SB_SESSION = requests.Session()


def _sb(method, path, data=None, headers=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    r = _SB_SESSION.request(method, url, json=data, headers=headers or _sb_headers(), timeout=25)
    r.raise_for_status()
    return r.json() if r.text else None


def _parse_fecha(v):
    if not v:
        return None
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


# ── Wompi ────────────────────────────────────────────────────────────────────
def _wompi_base():
    return ("https://sandbox.wompi.co/v1" if WOMPI_PUBLIC_KEY.startswith("pub_test_")
            else "https://production.wompi.co/v1")


def _firma(ref, cents, mon="COP"):
    return hashlib.sha256(f"{ref}{cents}{mon}{WOMPI_INTEGRITY_KEY}".encode()).hexdigest()


def _consultar_wompi(referencia="", transaction_id=""):
    base = _wompi_base()
    try:
        if transaction_id:
            r = requests.get(f"{base}/transactions/{urllib.parse.quote(transaction_id)}",
                             headers={"Authorization": f"Bearer {WOMPI_PUBLIC_KEY}"}, timeout=15)
            if r.status_code == 200:
                tx = r.json().get("data") or {}
                return tx if tx.get("status") == "APPROVED" else {}
            return {}
        if referencia and WOMPI_PRIVATE_KEY:
            r = requests.get(f"{base}/transactions",
                             headers={"Authorization": f"Bearer {WOMPI_PRIVATE_KEY}"},
                             params={"reference": referencia}, timeout=15)
            if r.status_code == 200:
                aprob = [t for t in (r.json().get("data") or []) if t.get("status") == "APPROVED"]
                aprob.sort(key=lambda x: x.get("created_at", ""), reverse=True)
                return aprob[0] if aprob else {}
        return {}
    except Exception:
        return {}


# ── SMTP ─────────────────────────────────────────────────────────────────────
def _enviar_correo(destino, asunto, texto, html=""):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = asunto
    msg["From"] = f"Lucreativity <{SMTP_REMITENTE}>"
    msg["To"] = destino
    msg.attach(MIMEText(texto, "plain"))
    if html:
        msg.attach(MIMEText(html, "html"))
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
        s.ehlo(); s.starttls(); s.ehlo()
        s.login(SMTP_REMITENTE, SMTP_APP_PWD)
        s.sendmail(SMTP_REMITENTE, destino, msg.as_string())


# ── Códigos OTP (guardados en Supabase, tabla codigos) ───────────────────────
def _guardar_codigo(clave, email, codigo):
    _sb("POST", "codigos?on_conflict=clave", data={
        "clave": clave, "email": email.lower(),
        "hash": hashlib.sha256(codigo.encode()).hexdigest(),
        "exp": (datetime.now() + timedelta(minutes=CODIGO_TTL_MIN)).isoformat(),
    }, headers=_sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"}))


def _validar_codigo(clave, email, codigo):
    res = _sb("GET", f"codigos?clave=eq.{urllib.parse.quote(clave)}&select=email,hash,exp")
    if not res:
        return "Primero solicita un código."
    d = res[0]
    # Supabase puede devolver 'exp' con zona horaria (timestamptz). Se compara de
    # forma segura: si la fecha es "aware", se usa un ahora también "aware".
    try:
        exp_dt = datetime.fromisoformat(str(d["exp"]).replace("Z", "+00:00"))
        ahora = datetime.now(exp_dt.tzinfo) if exp_dt.tzinfo else datetime.now()
        if exp_dt < ahora:
            return "El código expiró. Solicita uno nuevo."
    except Exception:
        pass  # si no se puede parsear la fecha, no se bloquea por expiración
    if email.lower() != d.get("email"):
        return "El correo no coincide con el del código."
    if hashlib.sha256(codigo.encode()).hexdigest() != d["hash"]:
        return "El código no es correcto."
    return ""


# ── Modelos ──────────────────────────────────────────────────────────────────
class EnviarCodigo(BaseModel):
    email: str

class Activar(BaseModel):
    serial: str
    device_id: str
    nombre_equipo: str = ""
    codigo: str
    cliente: dict

class Estado(BaseModel):
    serial: str

class Checkout(BaseModel):
    serial: str
    monto: int = 0

class VerificarPago(BaseModel):
    serial: str
    transaction_id: str = ""
    reference: str = ""

class EnviarCorreo(BaseModel):
    destino: str
    asunto: str
    texto: str
    html: str = ""

class Sync(BaseModel):
    tabla: str
    on_conflict: str = ""
    filas: list

class Restaurar(BaseModel):
    serial: str
    email: str
    device_id: str = ""
    nombre_equipo: str = ""


# ── Endpoints ────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"ok": True, "service": "pgirasa-backend"}


@app.post("/licencia/enviar-codigo")
def enviar_codigo(body: EnviarCodigo, x_app_key: str = Header(None)):
    _auth(x_app_key)
    email = body.email.strip()
    if "@" not in email:
        raise HTTPException(400, "Correo inválido.")
    codigo = f"{int.from_bytes(os.urandom(3), 'big') % 1000000:06d}"
    _guardar_codigo(f"act:{email.lower()}", email, codigo)
    try:
        _enviar_correo(email, "Código de activación - PGIRASA Tools",
                       f"Tu código de activación es: {codigo}\nVence en {CODIGO_TTL_MIN} minutos.",
                       f"<h2>Activación PGIRASA</h2><p style='font-size:26px;font-weight:bold'>{codigo}</p>")
    except Exception as e:
        raise HTTPException(502, f"No se pudo enviar el correo: {e}")
    return {"ok": True, "mensaje": "Código enviado al correo."}


@app.post("/licencia/estado")
def estado(body: Estado, x_app_key: str = Header(None)):
    _auth(x_app_key)
    res = _sb("GET", f"licencias?serial_compra=eq.{urllib.parse.quote(body.serial)}"
                     f"&select=fecha_vencimiento,tipo_licencia,precio,limite_pcs")
    if not res:
        return {"activa": False, "motivo": "La clave de licencia no existe."}
    lic = res[0]
    venc = _parse_fecha(lic.get("fecha_vencimiento"))
    if not venc:
        return {"activa": False, "motivo": "Licencia sin fecha de vencimiento."}
    dias = (venc - datetime.now().date()).days
    # La licencia se inactiva al llegar a 0 días: hay que renovar el pago.
    activa = dias > 0
    return {"activa": activa, "dias_restantes": dias,
            "vencimiento": venc.strftime("%Y-%m-%d"),
            "plan": lic.get("tipo_licencia") or "Mensual",
            "precio": int(lic.get("precio") or PAGO_MONTO),
            "motivo": "" if activa else "La licencia venció. Renueva el pago para continuar."}


def _actualizar_pcs(serial, total):
    """Refleja en licencias.pcs_registrados la cantidad de equipos registrados."""
    try:
        _sb("PATCH", f"licencias?serial_compra=eq.{urllib.parse.quote(serial)}",
            data={"pcs_registrados": int(total)},
            headers=_sb_headers({"Prefer": "return=minimal"}))
    except Exception:
        pass


@app.post("/licencia/activar")
def activar(body: Activar, x_app_key: str = Header(None)):
    _auth(x_app_key)
    cli = body.cliente or {}
    req_campos = ["razon_social", "nit", "direccion", "ciudad", "telefono", "email"]
    for c in req_campos:
        if not str(cli.get(c, "")).strip():
            raise HTTPException(400, f"Falta el dato obligatorio: {c}")
    err = _validar_codigo(f"act:{cli['email'].strip().lower()}", cli["email"], body.codigo)
    if err:
        return {"ok": False, "mensaje": err}
    res = _sb("GET", f"licencias?serial_compra=eq.{urllib.parse.quote(body.serial)}"
                     f"&select=serial_compra,fecha_vencimiento,limite_pcs")
    if not res:
        return {"ok": False, "mensaje": "La clave de licencia no es válida o no existe."}
    limite = int(res[0].get("limite_pcs") or 2)
    equipos = _sb("GET", f"equipos?serial_ref=eq.{urllib.parse.quote(body.serial)}&select=device_id") or []
    ids = [e.get("device_id") for e in equipos]
    cliente = {c: str(cli.get(c, "")).strip() for c in req_campos}
    cliente["serial_ref"] = body.serial
    _sb("POST", "clientes?on_conflict=serial_ref", data=cliente,
        headers=_sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"}))
    if body.device_id in ids:
        _actualizar_pcs(body.serial, len(ids))
        return {"ok": True, "mensaje": "Este equipo ya estaba registrado. Licencia activada."}
    if len(ids) >= limite:
        return {"ok": False, "mensaje": f"Esta licencia ya está en uso en {limite} equipos."}
    _sb("POST", "equipos", data={"serial_ref": body.serial, "device_id": body.device_id,
                                 "nombre_equipo": body.nombre_equipo,
                                 "registrado_en": datetime.now().isoformat()},
        headers=_sb_headers({"Prefer": "return=minimal"}))
    _actualizar_pcs(body.serial, len(ids) + 1)
    return {"ok": True, "mensaje": f"¡Licencia activada! Equipos en uso: {len(ids)+1} de {limite}."}


@app.post("/pago/checkout")
def checkout(body: Checkout, x_app_key: str = Header(None)):
    _auth(x_app_key)
    monto = body.monto if body.monto > 0 else PAGO_MONTO
    cents = monto * 100
    ref = f"PGIRASA-{body.serial}-{int(datetime.now().timestamp())}"
    params = {"public-key": WOMPI_PUBLIC_KEY, "currency": "COP",
              "amount-in-cents": str(cents), "reference": ref,
              "signature:integrity": _firma(ref, cents)}
    url = "https://checkout.wompi.co/p/?" + urllib.parse.urlencode(params)
    return {"url": url, "reference": ref}


@app.post("/pago/verificar")
def verificar(body: VerificarPago, x_app_key: str = Header(None)):
    _auth(x_app_key)
    tx = _consultar_wompi(transaction_id=body.transaction_id, referencia=body.reference)
    if not tx:
        return {"ok": False, "mensaje": "Aún no aparece un pago APROBADO en Wompi."}
    id_op = str(tx.get("id", ""))
    # Idempotencia: no aplicar dos veces el mismo pago
    if id_op and _sb("GET", f"pagos?id_operacion=eq.{urllib.parse.quote(id_op)}&select=id_operacion"):
        return {"ok": False, "mensaje": "Este pago ya había sido aplicado."}
    cust = tx.get("customer_data") or {}
    _sb("POST", "pagos?on_conflict=id_operacion", data={
        "id_operacion": id_op, "estado": "Aprobado",
        "monto": (tx.get("amount_in_cents", 0) or 0) / 100.0,
        "email_pagador": cust.get("email") or "", "serial_ref": body.serial,
        "fecha": tx.get("finalized_at") or tx.get("created_at") or datetime.now().isoformat(),
    }, headers=_sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"}))
    # Extender 30 días
    res = _sb("GET", f"licencias?serial_compra=eq.{urllib.parse.quote(body.serial)}&select=fecha_vencimiento")
    base = datetime.now().date()
    if res:
        v = _parse_fecha(res[0].get("fecha_vencimiento"))
        if v:
            base = max(base, v)
    nueva = (base + timedelta(days=30)).strftime("%Y-%m-%d")
    _sb("PATCH", f"licencias?serial_compra=eq.{urllib.parse.quote(body.serial)}",
        data={"fecha_vencimiento": nueva}, headers=_sb_headers({"Prefer": "return=minimal"}))
    return {"ok": True, "vencimiento": nueva, "mensaje": f"✓ Pago aprobado. Vence el {nueva}."}


@app.post("/email/enviar")
def email_enviar(body: EnviarCorreo, x_app_key: str = Header(None)):
    _auth(x_app_key)
    try:
        _enviar_correo(body.destino, body.asunto, body.texto, body.html)
    except Exception as e:
        raise HTTPException(502, f"No se pudo enviar el correo: {e}")
    return {"ok": True}


@app.post("/datos/sync")
def datos_sync(body: Sync, x_app_key: str = Header(None)):
    _auth(x_app_key)
    if not body.filas:
        return {"ok": True, "filas": 0}
    path = f"{body.tabla}?on_conflict={body.on_conflict}" if body.on_conflict else body.tabla
    try:
        _sb("POST", path, data=body.filas,
            headers=_sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"}))
    except requests.HTTPError as e:
        raise HTTPException(502, f"Error al sincronizar {body.tabla}: {e}")
    return {"ok": True, "filas": len(body.filas)}


# Tablas del negocio que se restauran al "Negocio Existente" (sin BLOBs/archivos).
TABLAS_RESTAURABLES = [
    "ajustes", "sedes", "usuarios", "rh1_diario", "entregas", "gastos",
    "presupuesto", "indicadores", "indicador_eval", "capacitaciones",
    "contingencias", "cronograma", "gagas_miembros", "gagas_reuniones",
    "auditoria",
]


@app.post("/datos/restaurar")
def datos_restaurar(body: Restaurar, x_app_key: str = Header(None)):
    """Restaura un negocio existente en un equipo nuevo.

    Valida la licencia y que el correo coincida con el registrado, registra el
    equipo (respetando el límite) y devuelve todos los datos del negocio para
    escribirlos en la base local.
    """
    _auth(x_app_key)
    serial = body.serial.strip()
    email = body.email.strip().lower()

    lic = _sb("GET", f"licencias?serial_compra=eq.{urllib.parse.quote(serial)}"
                     f"&select=serial_compra,limite_pcs")
    if not lic:
        return {"ok": False, "mensaje": "La clave de licencia no existe."}
    cli = _sb("GET", f"clientes?serial_ref=eq.{urllib.parse.quote(serial)}"
                     f"&select=razon_social,email,nit&limit=1")
    if not cli:
        return {"ok": False, "mensaje": "Esta licencia aún no tiene un negocio registrado."}
    if (cli[0].get("email") or "").strip().lower() != email:
        return {"ok": False, "mensaje": "El correo no coincide con el de la licencia."}

    # Registrar el equipo (respeta el límite de equipos de la licencia).
    if body.device_id:
        limite = int(lic[0].get("limite_pcs") or 2)
        equipos = _sb("GET", f"equipos?serial_ref=eq.{urllib.parse.quote(serial)}"
                             f"&select=device_id") or []
        ids = [e.get("device_id") for e in equipos]
        if body.device_id not in ids:
            if len(ids) >= limite:
                return {"ok": False,
                        "mensaje": f"Esta licencia ya está en uso en {limite} equipos."}
            _sb("POST", "equipos",
                data={"serial_ref": serial, "device_id": body.device_id,
                      "nombre_equipo": body.nombre_equipo,
                      "registrado_en": datetime.now().isoformat()},
                headers=_sb_headers({"Prefer": "return=minimal"}))
            _actualizar_pcs(serial, len(ids) + 1)
        else:
            _actualizar_pcs(serial, len(ids))

    # Descargar SOLO las tablas de esta empresa (filtradas por su NIT).
    nit_emp = (cli[0].get("nit") or "").strip()
    nit_q = urllib.parse.quote(nit_emp)
    tablas = {}
    for t in TABLAS_RESTAURABLES:
        try:
            tablas[t] = _sb("GET", f"{t}?nit=eq.{nit_q}&select=*") or []
        except Exception:
            tablas[t] = []
    return {"ok": True, "entidad": cli[0].get("razon_social", ""),
            "nit": nit_emp, "tablas": tablas}


# ============================================================
#  APP MÓVIL (PGIRASAtoolsMobile)
#  La app móvil solo habla con estos endpoints (sin llaves locales).
# ============================================================
TOKEN_TTL_DIAS = 30

# Categorías de residuos RH1 (interna, peligroso?)
RESIDUOS = [
    ("no_bolsa_blanca", False), ("no_bolsa_negra", False), ("no_bolsa_verde", False),
    ("pel_biosanitario", True), ("pel_anatomopatologico", True), ("pel_cortopunzante", True),
    ("pel_animales", True), ("pel_farmacos", True), ("pel_citotoxicos", True),
    ("pel_metales", True), ("pel_reactivos", True), ("pel_contenedores", True),
    ("pel_aceites", True), ("pel_radiactivo", True), ("ap_raee", False),
]
RESIDUO_COLS = [c[0] for c in RESIDUOS]


def _crear_token(username, role, sede, nit):
    exp = int(time.time()) + TOKEN_TTL_DIAS * 86400
    payload = f"{username}|{role}|{sede}|{nit}|{exp}"
    sig = hmac.new(APP_API_KEY.encode(), payload.encode(), hashlib.sha256).hexdigest()
    return base64.urlsafe_b64encode(payload.encode()).decode() + "." + sig


def _validar_token(token):
    try:
        b64, sig = token.split(".", 1)
        payload = base64.urlsafe_b64decode(b64.encode()).decode()
        esperado = hmac.new(APP_API_KEY.encode(), payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, esperado):
            return None
        partes = payload.split("|")
        if len(partes) == 5:        # nuevo formato: incluye nit
            username, role, sede, nit, exp = partes
        elif len(partes) == 4:      # formato anterior (sin nit): forzar re-login
            return None
        else:
            return None
        if int(exp) < int(time.time()):
            return None
        return {"username": username, "role": role, "sede": sede, "nit": nit}
    except Exception:
        return None


def _verificar_pwd(pwd, stored):
    if not stored:
        return False
    if stored.startswith("$2"):
        try:
            return bcrypt.checkpw(pwd.encode(), stored.encode())
        except Exception:
            return False
    return stored == hashlib.sha256(pwd.encode()).hexdigest()


class MovilLogin(BaseModel):
    nit: str
    usuario: str
    password: str

class MovilRegistrar(BaseModel):
    token: str
    sede: str
    fecha: str
    valores: dict
    observaciones: str = ""

class MovilToken(BaseModel):
    token: str

class MovilBloquear(BaseModel):
    token: str
    username: str
    bloquear: bool


@app.post("/movil/login")
def movil_login(body: MovilLogin, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    from concurrent.futures import ThreadPoolExecutor
    usuario = body.usuario.strip()
    if _login_bloqueado(usuario):
        return {"ok": False, "mensaje": "Demasiados intentos fallidos. "
                "Espera unos minutos e inténtalo de nuevo."}
    nit = body.nit.strip()
    # Las 3 consultas (cliente por NIT, usuario y sedes) son independientes:
    # se lanzan en paralelo. Devolver las sedes en el login evita una consulta
    # extra después (acelera la pantalla de registro del usuario).
    with ThreadPoolExecutor(max_workers=3) as ex:
        f_cli = ex.submit(_sb, "GET",
                          f"clientes?nit=eq.{urllib.parse.quote(nit)}&select=razon_social&limit=1")
        f_usr = ex.submit(_sb, "GET",
                          f"usuarios?nit=eq.{urllib.parse.quote(nit)}"
                          f"&username=eq.{urllib.parse.quote(usuario)}"
                          f"&select=username,password_hash,nombre_completo,rol,activo,"
                          f"sede_asignada,bloqueado")
        f_sed = ex.submit(_sb, "GET",
                          f"sedes?nit=eq.{urllib.parse.quote(nit)}"
                          f"&activo=eq.1&select=nombre&order=nombre")
        cli = f_cli.result()
        res = f_usr.result()
        try:
            sedes_all = f_sed.result() or []
        except Exception:
            sedes_all = []
    if not cli:
        return {"ok": False, "mensaje": "NIT no encontrado."}
    if not res:
        _login_fallo(usuario)
        return {"ok": False, "mensaje": "Usuario o contraseña incorrectos."}
    u = res[0]
    if not u.get("activo", 1):
        return {"ok": False, "mensaje": "La cuenta está inactiva."}
    if u.get("bloqueado"):
        return {"ok": False, "mensaje": "Cuenta bloqueada. Contacta al administrador."}
    if not _verificar_pwd(body.password, u.get("password_hash", "")):
        _login_fallo(usuario)
        return {"ok": False, "mensaje": "Usuario o contraseña incorrectos."}
    _login_ok(usuario)
    sede = u.get("sede_asignada") or ""
    token = _crear_token(u["username"], u.get("rol", "usuario"), sede, nit)
    sedes = [s["nombre"] for s in sedes_all if s.get("nombre")]
    if sede:  # usuario con sede asignada: solo la suya
        sedes = [x for x in sedes if x == sede]
    return {"ok": True, "token": token, "rol": u.get("rol", "usuario"),
            "nombre": u.get("nombre_completo", ""), "sede_asignada": sede,
            "entidad": cli[0].get("razon_social", ""), "sedes": sedes}


@app.post("/movil/sedes")
def movil_sedes(body: MovilToken, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t:
        raise HTTPException(401, "Sesión expirada. Inicia sesión de nuevo.")
    nit = urllib.parse.quote(t["nit"])
    res = _sb("GET", f"sedes?nit=eq.{nit}&activo=eq.1&select=nombre&order=nombre") or []
    todas = [r["nombre"] for r in res]
    if t["sede"]:
        todas = [s for s in todas if s == t["sede"]]
    return {"ok": True, "sedes": todas}


@app.post("/movil/registrar")
def movil_registrar(body: MovilRegistrar, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t:
        raise HTTPException(401, "Sesión expirada. Inicia sesión de nuevo.")
    if t["sede"] and t["sede"] != body.sede:
        return {"ok": False, "mensaje": "No tienes permiso para registrar en esa sede."}
    try:
        dt = datetime.strptime(body.fecha, "%Y-%m-%d")
    except Exception:
        return {"ok": False, "mensaje": "Fecha inválida."}
    fila = {"nit": t["nit"], "fecha": body.fecha, "anio": dt.year, "mes": dt.month,
            "dia": dt.day, "sede": body.sede, "responsable": t["username"],
            "observaciones": body.observaciones}
    # rh1_diario usa PK (nit, id) sin default; el id se deriva de (fecha, sede)
    # para mantener el mismo registro al regrabar el día/sede.
    fila["id"] = 1_000_000_000 + int.from_bytes(
        hashlib.sha1(f"{body.fecha}|{body.sede}".encode()).digest()[:6], "big")
    for c in RESIDUO_COLS:
        fila[c] = float(body.valores.get(c, 0) or 0)
    try:
        _sb("POST", "rh1_diario?on_conflict=nit,fecha,sede", data=fila,
            headers=_sb_headers({"Prefer": "resolution=merge-duplicates,return=minimal"}))
    except requests.HTTPError as e:
        raise HTTPException(502, f"No se pudo guardar: {e}")
    return {"ok": True, "mensaje": f"Residuos del {body.fecha} guardados para {body.sede}."}


@app.post("/movil/estadisticas")
def movil_estadisticas(body: MovilToken, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t or t["role"] != "admin":
        raise HTTPException(403, "Solo administradores.")
    anio = datetime.now().year
    nit = urllib.parse.quote(t["nit"])
    sel = ",".join(["sede", "fecha"] + RESIDUO_COLS)
    rows = _sb("GET", f"rh1_diario?nit=eq.{nit}&anio=eq.{anio}&select={sel}") or []
    por_sede = {}
    total_pel = total_nopel = 0.0
    pelig = {c for c, p in RESIDUOS if p}
    for r in rows:
        s = r.get("sede", "") or "(sin sede)"
        acc = por_sede.setdefault(s, {"peligrosos": 0.0, "no_peligrosos": 0.0, "registros": 0})
        acc["registros"] += 1
        for c in RESIDUO_COLS:
            v = float(r.get(c, 0) or 0)
            if c in pelig:
                acc["peligrosos"] += v; total_pel += v
            else:
                acc["no_peligrosos"] += v; total_nopel += v
    return {"ok": True, "anio": anio, "por_sede": por_sede,
            "total_peligrosos": round(total_pel, 2),
            "total_no_peligrosos": round(total_nopel, 2),
            "total_general": round(total_pel + total_nopel, 2)}


@app.post("/movil/usuarios")
def movil_usuarios(body: MovilToken, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t or t["role"] != "admin":
        raise HTTPException(403, "Solo administradores.")
    nit = urllib.parse.quote(t["nit"])
    res = _sb("GET", f"usuarios?nit=eq.{nit}"
                     "&select=username,nombre_completo,rol,activo,sede_asignada,bloqueado"
                     "&order=username") or []
    return {"ok": True, "usuarios": res}


@app.post("/movil/bloquear")
def movil_bloquear(body: MovilBloquear, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t or t["role"] != "admin":
        raise HTTPException(403, "Solo administradores.")
    nit = urllib.parse.quote(t["nit"])
    _sb("PATCH", f"usuarios?nit=eq.{nit}&username=eq.{urllib.parse.quote(body.username)}",
        data={"bloqueado": 1 if body.bloquear else 0},
        headers=_sb_headers({"Prefer": "return=minimal"}))
    return {"ok": True, "mensaje": ("Usuario bloqueado." if body.bloquear else "Usuario desbloqueado.")}


# ============================================================
#  DASHBOARD Y REPORTES MÓVIL (solo administradores)
#  Calcula sobre Supabase lo que el escritorio calcula en local:
#  estadísticas, tendencias, semáforo de cumplimiento e indicadores.
# ============================================================
MESES_ES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
_PELIGROSOS = {c for c, p in RESIDUOS if p}


def _num(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def _rh1_anio(anio):
    sel = ",".join(["sede", "fecha", "mes"] + RESIDUO_COLS)
    return _sb("GET", f"rh1_diario?anio=eq.{anio}&select={sel}") or []


def _estadisticas(rows):
    por_sede, tot_pel, tot_no = {}, 0.0, 0.0
    for r in rows:
        s = r.get("sede") or "(sin sede)"
        acc = por_sede.setdefault(s, {"peligrosos": 0.0, "no_peligrosos": 0.0, "registros": 0})
        acc["registros"] += 1
        for c in RESIDUO_COLS:
            v = _num(r.get(c))
            if c in _PELIGROSOS:
                acc["peligrosos"] += v; tot_pel += v
            else:
                acc["no_peligrosos"] += v; tot_no += v
    for d in por_sede.values():
        d["peligrosos"] = round(d["peligrosos"], 2)
        d["no_peligrosos"] = round(d["no_peligrosos"], 2)
    return {"por_sede": por_sede, "total_peligrosos": round(tot_pel, 2),
            "total_no_peligrosos": round(tot_no, 2),
            "total_general": round(tot_pel + tot_no, 2)}


def _tendencias(rows, anio, mes):
    por_mes = {}
    for r in rows:
        m = int(r.get("mes") or 0)
        por_mes[m] = por_mes.get(m, 0.0) + sum(_num(r.get(c)) for c in RESIDUO_COLS)
    total_anio = round(sum(por_mes.values()), 2)
    total_mes = round(por_mes.get(mes, 0.0), 2)
    mes_ant = round(por_mes.get(mes - 1, 0.0), 2) if mes > 1 else 0.0
    var = round((total_mes - mes_ant) / mes_ant * 100, 1) if mes_ant else None
    con_datos = [por_mes[m] for m in range(1, mes + 1) if por_mes.get(m, 0) > 0]
    proy = round(sum(con_datos) / len(con_datos) * 12, 1) if con_datos else 0.0
    return {"total_anio": total_anio, "total_mes": total_mes,
            "mes_anterior": mes_ant, "var_mes_pct": var,
            "proyeccion_anual": proy, "mes_nombre": MESES_ES[mes]}


def _indicadores_from(inds, evals):
    """Evalúa indicadores a partir de datos ya consultados (sin N+1).

    `evals` es la lista completa de indicador_eval ordenada por fecha desc; se
    toma la primera ocurrencia de cada indicador (la más reciente).
    """
    ultimo = {}
    for e in evals or []:
        iid = e.get("indicador_id")
        if iid not in ultimo:
            ultimo[iid] = e
    out, incum, sin_eval = [], 0, 0
    for ind in inds or []:
        base = {"codigo": ind.get("codigo", ""), "nombre": ind.get("nombre", ""),
                "unidad": ind.get("unidad", "%"), "meta": _num(ind.get("meta"))}
        e = ultimo.get(ind.get("id"))
        if e:
            cumple = bool(e.get("cumple"))
            if not cumple:
                incum += 1
            out.append({**base, "resultado": round(_num(e.get("resultado")), 2),
                        "periodo": e.get("periodo") or e.get("fecha", ""), "cumple": cumple})
        else:
            sin_eval += 1
            out.append({**base, "resultado": None, "periodo": "", "cumple": None})
    return out, incum, sin_eval


def _semaforo_from(rows, anio, mes, ind_stats, entregas, cronograma,
                   contingencias, gagas, presupuesto, gastos):
    """Semáforo de cumplimiento a partir de datos ya consultados (sin red)."""
    out = []
    hoy = datetime.now().date()

    def add(modulo, estado, detalle):
        out.append({"modulo": modulo, "estado": estado, "detalle": detalle})

    try:  # RH1: días sin registrar este mes
        fechas = {r["fecha"] for r in rows if int(r.get("mes") or 0) == mes}
        falt = sum(1 for d in range(1, hoy.day + 1)
                   if f"{anio}-{mes:02d}-{d:02d}" not in fechas)
        if falt == 0:
            add("RH1", "verde", "Registro diario al día")
        elif falt <= 3:
            add("RH1", "amarillo", f"{falt} día(s) sin registrar este mes")
        else:
            add("RH1", "rojo", f"{falt} días sin registrar este mes")
    except Exception:
        pass

    try:  # Entregas sin certificado del gestor (año)
        n = sum(1 for e in (entregas or []) if not (e.get("num_certificado") or "").strip())
        if n == 0:
            add("Entregas", "verde", "Todas las entregas con certificado")
        elif n <= 2:
            add("Entregas", "amarillo", f"{n} entrega(s) sin certificado")
        else:
            add("Entregas", "rojo", f"{n} entregas sin certificado")
    except Exception:
        pass

    try:  # Indicadores (ya calculado)
        incum, sin_eval = ind_stats
        total = incum + sin_eval
        if incum == 0 and sin_eval == 0:
            add("Indicadores", "verde", "Todos cumplen su meta")
        elif incum >= 2 or (total and sin_eval == total):
            add("Indicadores", "rojo", f"{incum} sin cumplir, {sin_eval} sin evaluar")
        else:
            add("Indicadores", "amarillo", f"{incum} sin cumplir, {sin_eval} sin evaluar")
    except Exception:
        pass

    try:  # Cronograma vencido
        venc = sum(1 for a in (cronograma or [])
                   if a.get("estado") not in ("Completada", "Cancelada")
                   and a.get("plazo_fecha") and a["plazo_fecha"] < hoy.isoformat())
        if venc == 0:
            add("Cronograma", "verde", "Sin actividades vencidas")
        elif venc <= 2:
            add("Cronograma", "amarillo", f"{venc} actividad(es) vencida(s)")
        else:
            add("Cronograma", "rojo", f"{venc} actividades vencidas")
    except Exception:
        pass

    try:  # Contingencias abiertas
        ab = sum(1 for c in (contingencias or []) if c.get("estado") != "Cerrada")
        if ab == 0:
            add("Contingencias", "verde", "Sin contingencias abiertas")
        elif ab <= 1:
            add("Contingencias", "amarillo", "1 contingencia abierta")
        else:
            add("Contingencias", "rojo", f"{ab} contingencias abiertas")
    except Exception:
        pass

    try:  # GAGAS última reunión
        fechas_g = [r.get("fecha", "") for r in (gagas or []) if r.get("fecha")]
        ultima = max(fechas_g) if fechas_g else ""
        lim_a = (hoy - timedelta(days=60)).isoformat()
        lim_r = (hoy - timedelta(days=90)).isoformat()
        if ultima and ultima >= lim_a:
            add("GAGAS", "verde", f"Última reunión: {ultima}")
        elif ultima and ultima >= lim_r:
            add("GAGAS", "amarillo", f"Última reunión: {ultima}")
        else:
            add("GAGAS", "rojo", f"Última reunión: {ultima or 'sin registros'}")
    except Exception:
        pass

    try:  # Presupuesto
        ejec = {}
        for g in (gastos or []):
            ejec[g.get("rubro_id")] = ejec.get(g.get("rubro_id"), 0) + _num(g.get("valor"))
        sobre = [p for p in (presupuesto or []) if _num(p.get("monto_asignado")) > 0
                 and ejec.get(p["id"], 0) > _num(p.get("monto_asignado"))]
        casi = [p for p in (presupuesto or []) if _num(p.get("monto_asignado")) > 0
                and 0.9 <= ejec.get(p["id"], 0) / _num(p.get("monto_asignado")) <= 1]
        if sobre:
            add("Presupuesto", "rojo", f"{len(sobre)} rubro(s) sobreejecutado(s)")
        elif casi:
            add("Presupuesto", "amarillo", f"{len(casi)} rubro(s) ≥90% ejecutado(s)")
        elif presupuesto:
            add("Presupuesto", "verde", "Ejecución bajo control")
    except Exception:
        pass

    return out


def _dashboard_data(t):
    """Reúne todos los datos del panel haciendo las consultas en PARALELO.

    Antes se hacían ~20 consultas secuenciales (incluida una por indicador);
    ahora son ~11 en paralelo y los indicadores en una sola consulta.
    """
    from concurrent.futures import ThreadPoolExecutor

    anio = datetime.now().year
    mes = datetime.now().month
    nit = urllib.parse.quote(t["nit"])
    sel_rh1 = ",".join(["sede", "fecha", "mes"] + RESIDUO_COLS)

    pedidos = {
        "rows": f"rh1_diario?nit=eq.{nit}&anio=eq.{anio}&select={sel_rh1}",
        "inds": f"indicadores?nit=eq.{nit}&select=id,codigo,nombre,unidad,meta,frecuencia&order=codigo",
        "evals": f"indicador_eval?nit=eq.{nit}&select=indicador_id,fecha,periodo,resultado,cumple&order=fecha.desc",
        "entregas": f"entregas?nit=eq.{nit}&fecha=like.{anio}-*&select=num_certificado",
        "cronograma": f"cronograma?nit=eq.{nit}&anio=eq.{anio}&select=estado,plazo_fecha",
        "contingencias": f"contingencias?nit=eq.{nit}&select=estado",
        "gagas": f"gagas_reuniones?nit=eq.{nit}&select=fecha&order=fecha.desc&limit=1",
        "presupuesto": f"presupuesto?nit=eq.{nit}&anio=eq.{anio}&select=id,monto_asignado",
        "gastos": f"gastos?nit=eq.{nit}&anio=eq.{anio}&select=rubro_id,valor",
        "usuarios": (f"usuarios?nit=eq.{nit}&select=username,nombre_completo,rol,activo,"
                     "sede_asignada,bloqueado&order=username"),
        "clientes": f"clientes?nit=eq.{urllib.parse.quote(t['nit'])}&select=razon_social&limit=1",
    }

    def g(path):
        try:
            return _sb("GET", path) or []
        except Exception:
            return []

    with ThreadPoolExecutor(max_workers=8) as ex:
        fut = {k: ex.submit(g, p) for k, p in pedidos.items()}
        D = {k: f.result() for k, f in fut.items()}

    rows = D["rows"]
    indicadores, incum, sin_eval = _indicadores_from(D["inds"], D["evals"])
    entidad = D["clientes"][0].get("razon_social", "") if D["clientes"] else ""
    return {
        "anio": anio,
        "estadisticas": _estadisticas(rows),
        "tendencias": _tendencias(rows, anio, mes),
        "semaforo": _semaforo_from(rows, anio, mes, (incum, sin_eval),
                                   D["entregas"], D["cronograma"], D["contingencias"],
                                   D["gagas"], D["presupuesto"], D["gastos"]),
        "indicadores": indicadores,
        "entidad": entidad,
        "usuarios": D["usuarios"],
    }


@app.post("/movil/dashboard")
def movil_dashboard(body: MovilToken, x_app_key: str = Header(None)):
    _auth_movil(x_app_key)
    t = _validar_token(body.token)
    if not t or t["role"] != "admin":
        raise HTTPException(403, "Solo administradores.")
    return {"ok": True, **_dashboard_data(t)}


# ── Generación de informes (PDF / Excel) ─────────────────────────────────────
def _informe_pdf(data, entidad):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.6 * cm,
                            bottomMargin=1.6 * cm, leftMargin=1.6 * cm, rightMargin=1.6 * cm)
    st = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=st["Title"], textColor=colors.HexColor("#1B4E73"), fontSize=20)
    h2 = ParagraphStyle("h2", parent=st["Heading2"], textColor=colors.HexColor("#2E7FB8"))
    small = ParagraphStyle("small", parent=st["Normal"], fontSize=9, textColor=colors.HexColor("#666666"))
    AZUL = colors.HexColor("#2E7FB8")
    el = []

    el.append(Paragraph("Informe General PGIRASA", h1))
    if entidad or data.get("entidad"):
        el.append(Paragraph(entidad or data.get("entidad", ""), st["Heading3"]))
    el.append(Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} · "
                        f"Año {data['anio']}", small))
    el.append(Spacer(1, 0.4 * cm))

    est = data["estadisticas"]
    ten = data["tendencias"]
    el.append(Paragraph("Resumen de residuos", h2))
    resumen = [["Indicador", "Valor (kg)"],
               ["Peligrosos", f"{est['total_peligrosos']:,.1f}"],
               ["No peligrosos", f"{est['total_no_peligrosos']:,.1f}"],
               ["Total año", f"{est['total_general']:,.1f}"],
               [f"Total {ten['mes_nombre']}", f"{ten['total_mes']:,.1f}"],
               ["Proyección anual", f"{ten['proyeccion_anual']:,.1f}"]]
    t1 = Table(resumen, colWidths=[8 * cm, 6 * cm])
    t1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CAD6E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F7")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9)]))
    el.append(t1)
    el.append(Spacer(1, 0.4 * cm))

    el.append(Paragraph("Generación por sede", h2))
    filas = [["Sede", "Peligrosos", "No pelig.", "Registros"]]
    for s, d in est["por_sede"].items():
        filas.append([s, f"{d['peligrosos']:,.1f}", f"{d['no_peligrosos']:,.1f}", str(d["registros"])])
    if len(filas) == 1:
        filas.append(["(sin datos)", "0", "0", "0"])
    t2 = Table(filas, colWidths=[6 * cm, 3 * cm, 3 * cm, 2.5 * cm])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CAD6E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F7")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9)]))
    el.append(t2)
    el.append(Spacer(1, 0.4 * cm))

    el.append(Paragraph("Semáforo de cumplimiento", h2))
    _col = {"verde": colors.HexColor("#2E7D32"), "amarillo": colors.HexColor("#B77900"),
            "rojo": colors.HexColor("#C62828")}
    sfilas = [["Módulo", "Estado", "Detalle"]]
    for s in data["semaforo"]:
        sfilas.append([s["modulo"], s["estado"].upper(), s["detalle"]])
    if len(sfilas) == 1:
        sfilas.append(["(sin datos)", "-", ""])
    t3 = Table(sfilas, colWidths=[4 * cm, 2.5 * cm, 8 * cm])
    estilo3 = [("BACKGROUND", (0, 0), (-1, 0), AZUL),
               ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
               ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CAD6E5")),
               ("FONTSIZE", (0, 0), (-1, -1), 9)]
    for i, s in enumerate(data["semaforo"], start=1):
        estilo3.append(("TEXTCOLOR", (1, i), (1, i), _col.get(s["estado"], colors.black)))
    t3.setStyle(TableStyle(estilo3))
    el.append(t3)
    el.append(Spacer(1, 0.4 * cm))

    el.append(Paragraph("Evaluación de indicadores (Res. 591/2024)", h2))
    ifilas = [["Código", "Indicador", "Meta", "Resultado", "Cumple"]]
    for ind in data["indicadores"]:
        res = "—" if ind["resultado"] is None else f"{ind['resultado']:,.1f}{ind['unidad']}"
        cum = "Sin evaluar" if ind["cumple"] is None else ("Sí" if ind["cumple"] else "No")
        ifilas.append([ind["codigo"], ind["nombre"][:40],
                       f"{ind['meta']:,.0f}{ind['unidad']}", res, cum])
    if len(ifilas) == 1:
        ifilas.append(["—", "(sin indicadores)", "—", "—", "—"])
    t4 = Table(ifilas, colWidths=[2 * cm, 6.5 * cm, 2.2 * cm, 2.5 * cm, 2 * cm])
    t4.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CAD6E5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F4F7")]),
        ("FONTSIZE", (0, 0), (-1, -1), 8)]))
    el.append(t4)

    doc.build(el)
    return buf.getvalue(), "application/pdf", f"informe_pgirasa_{data['anio']}.pdf"


def _informe_xlsx(data, entidad):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    azul = PatternFill("solid", fgColor="2E7FB8")
    blanco = Font(color="FFFFFF", bold=True)
    cen = Alignment(horizontal="center")

    def encabezado(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.fill = azul; c.font = blanco; c.alignment = cen

    est = data["estadisticas"]; ten = data["tendencias"]
    ws = wb.active; ws.title = "Resumen"
    encabezado(ws, ["Indicador", "Valor (kg)"])
    for k, v in [("Peligrosos", est["total_peligrosos"]),
                 ("No peligrosos", est["total_no_peligrosos"]),
                 ("Total año", est["total_general"]),
                 (f"Total {ten['mes_nombre']}", ten["total_mes"]),
                 ("Mes anterior", ten["mes_anterior"]),
                 ("Proyección anual", ten["proyeccion_anual"])]:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 26; ws.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("Por sede")
    encabezado(ws2, ["Sede", "Peligrosos", "No peligrosos", "Registros"])
    for s, d in est["por_sede"].items():
        ws2.append([s, d["peligrosos"], d["no_peligrosos"], d["registros"]])
    for col, an in zip("ABCD", (26, 14, 14, 12)):
        ws2.column_dimensions[col].width = an

    ws3 = wb.create_sheet("Semáforo")
    encabezado(ws3, ["Módulo", "Estado", "Detalle"])
    for s in data["semaforo"]:
        ws3.append([s["modulo"], s["estado"].upper(), s["detalle"]])
    for col, an in zip("ABC", (20, 14, 50)):
        ws3.column_dimensions[col].width = an

    ws4 = wb.create_sheet("Indicadores")
    encabezado(ws4, ["Código", "Indicador", "Meta", "Unidad", "Resultado", "Periodo", "Cumple"])
    for ind in data["indicadores"]:
        ws4.append([ind["codigo"], ind["nombre"], ind["meta"], ind["unidad"],
                    ind["resultado"], ind["periodo"],
                    "Sin evaluar" if ind["cumple"] is None else ("Sí" if ind["cumple"] else "No")])
    for col, an in zip("ABCDEFG", (12, 40, 10, 8, 12, 14, 12)):
        ws4.column_dimensions[col].width = an

    buf = BytesIO(); wb.save(buf)
    return (buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"informe_pgirasa_{data['anio']}.xlsx")


@app.get("/movil/informe")
def movil_informe(token: str, formato: str = "pdf", entidad: str = ""):
    """Descarga directa del informe (autenticado por el token firmado de la app).

    Se accede por GET para poder abrirlo desde el navegador del celular.
    """
    t = _validar_token(token)
    if not t or t["role"] != "admin":
        raise HTTPException(403, "Solo administradores.")
    data = _dashboard_data(t)
    try:
        if (formato or "pdf").lower() == "xlsx":
            contenido, mime, fn = _informe_xlsx(data, entidad)
        else:
            contenido, mime, fn = _informe_pdf(data, entidad)
    except ImportError as e:
        raise HTTPException(500, f"Falta una dependencia en el backend: {e}. "
                                 "Agrega 'reportlab' y 'openpyxl' a requirements.txt.")
    return Response(content=contenido, media_type=mime,
                    headers={"Content-Disposition": f'attachment; filename="{fn}"'})
